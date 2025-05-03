import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime
import os
import pandas as pd



def split_string_by_length(text, N):
    wtext   = text.replace('\n', "")
    splits  = wtext.split(" ")
    lines   = []
    result  = {}
    start   = 0

    limiter = 0
    carrier = []
    for word in splits:
        carrier.append(word)
        limiter += 1
        if limiter > N:
            lines.append(carrier)
            limiter = 0
            carrier = []

    lines.append(carrier)
    
    for i, line in enumerate(lines[:3]):
        sline = " ".join(line)
        result[i] = sline

    for i in range(3):
        if i not in result:
            result[i] = ""
    
    return result



def merge_cells_by_row_col(sheet, start_row, start_col, end_row, end_col):
  """
  Merges cells in an openpyxl sheet using row and column numbers.

  Args:
    sheet: The openpyxl worksheet object.
    start_row: The row number of the top-left cell of the merge range (1-indexed).
    start_col: The column number of the top-left cell of the merge range (1-indexed).
    end_row: The row number of the bottom-right cell of the merge range (1-indexed).
    end_col: The column number of the bottom-right cell of the merge range (1-indexed).
  """

  # Construct the cell range string in the format "A1:B2"
  start_cell = sheet.cell(row=start_row, column=start_col).coordinate
  end_cell = sheet.cell(row=end_row, column=end_col).coordinate
  merge_range = f"{start_cell}:{end_cell}"

  # Merge the cells
  sheet.merge_cells(merge_range)



def generate_invoice(writer, invoice_data):
    workbook  = writer.book
    worksheet = workbook.add_worksheet('Invoice')
    
    # Define formats
    title_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 16,
        'bold': True,
        'align': 'center'
    })
    
    header_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 12,
        'bold': True
    })
    
    normal_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 12
    })
    
    header_fill_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 12,
        'bold': True,
        'bg_color': 'FFD1DC',
        'border': 1,
        'border_color': '4C2882',
        'align': 'center',
        'text_wrap': True
    })
    
    thin_border_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 12,
        'border': 1,
        'border_color': '4C2882'
    })
    
    right_aligned_border_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 12,
        'border': 1,
        'border_color': '4C2882',
        'align': 'right'
    })
    
    header_border_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 12,
        'bold': True,
        'border': 1,
        'border_color': '4C2882'
    })
    
    header_left_border_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 12,
        'bold': True,
        'border': 1,
        'border_color': '4C2882',
        'align': 'left'
    })
    
    header_right_border_format = workbook.add_format({
        'font_name': 'Times New Roman',
        'font_size': 12,
        'bold': True,
        'border': 1,
        'border_color': '4C2882',
        'align': 'right'
    })
    
    # Set column widths
    worksheet.set_column('A:A', 5)
    worksheet.set_column('B:B', 40)
    worksheet.set_column('C:C', 15)
    worksheet.set_column('D:D', 15)
    worksheet.set_column('E:E', 15)
    worksheet.set_column('F:F', 20)
    worksheet.set_column('G:G', 5)
    
    # Hide gridlines
    worksheet.hide_gridlines(2)  # Hide screen gridlines
    
    # Title
    worksheet.merge_range('B3:F3', 'INVOICE', title_format)
    
    # Customer information
    worksheet.write('B5', 'Ditagihkan kepada:', normal_format)
    worksheet.write('B6', invoice_data['customer_name'], header_format)
    worksheet.write('B7', invoice_data['customer_address_1'], normal_format)
    worksheet.write('B8', invoice_data['customer_address_2'], normal_format)
    worksheet.write('B9', invoice_data['customer_address_3'], normal_format)
    
    # Company information
    worksheet.write('E6', 'No. Invoice', normal_format)
    worksheet.write('E7', 'ID Pelanggan', normal_format)
    worksheet.write('E8', 'Periode', normal_format)
    worksheet.write('E9', 'Tgl. Invoice', normal_format)
    worksheet.write('E10', 'Tgl. Jatuh Tempo', normal_format)
    
    worksheet.write('F6', f": {invoice_data['invoice_number']}", normal_format)
    worksheet.write('F7', f": {invoice_data['customer_id']}", normal_format)
    worksheet.write('F8', f": {invoice_data['invoice_period']}", normal_format)
    worksheet.write('F9', f": {invoice_data['date']}", normal_format)
    worksheet.write('F10', f": {invoice_data['due_date']}", normal_format)
    
    # Invoice items header
    headers = ['Item', 'Volume      (Sm3)', 'Harga            (Rp/Sm3)', 'Nilai         Tagihan (Rp)', 'Keterangan']
    for col, header in enumerate(headers):
        worksheet.write(12, col + 1, header, header_fill_format)
    
    # Invoice items
    row = 13
    for item in invoice_data['items']:
        worksheet.write(row, 1, item['item'], thin_border_format)
        worksheet.write(row, 2, item['volume'], right_aligned_border_format)
        worksheet.write(row, 3, item['unit_price'], right_aligned_border_format)
        worksheet.write(row, 4, item['price'], right_aligned_border_format)
        worksheet.write(row, 5, item['note'], thin_border_format)
        row += 1
    
    conclusion_row = row
    dpp_row = row + 1
    ppn_row = row + 2
    total_row = row + 3
    
    # Conclusion row
    inweek  = invoice_data['inweek']
    inmonth = invoice_data['inmonth']
    inyear  = invoice_data['inyear']
    inprice = invoice_data['inprice']
    
    worksheet.merge_range(conclusion_row, 1, conclusion_row, 3, 
                         f"Pemakaian Gas CNG Periode {invoice_data["invoice_period"]} )*", 
                         header_border_format)
    worksheet.write(conclusion_row, 4, f"{inprice}", header_right_border_format)
    worksheet.write(conclusion_row, 5, "", thin_border_format)  # For the note column border
    
    # DPP
    worksheet.write(dpp_row, 3, "DPP  )**", header_left_border_format)
    worksheet.write(dpp_row, 4, f"{invoice_data['dpp_price']}", header_right_border_format)
    worksheet.write(dpp_row, 5, "", thin_border_format)  # For the note column border
    
    # PPN
    worksheet.write(ppn_row, 3, "PPN", header_left_border_format)
    worksheet.write(ppn_row, 4, f"{invoice_data['charged_tax']}", header_right_border_format)
    worksheet.write(ppn_row, 5, "", thin_border_format)  # For the note column border
    
    # TOTAL
    worksheet.write(total_row, 3, "TOTAL", header_left_border_format)
    worksheet.write(total_row, 4, f"{invoice_data['total_taxed']}", header_right_border_format)
    worksheet.write(total_row, 5, "", thin_border_format)  # For the note column border
    
    # Signature
    sign_row = total_row + 3
    worksheet.write(sign_row, 4, "Hormat kami,", normal_format)
    worksheet.write(sign_row + 4, 4, "Alice Alisceon", normal_format)
    worksheet.write(sign_row + 5, 4, "Direktur", normal_format)
    
    # Payment terms and notes
    notes_row = sign_row + 10
    worksheet.write(notes_row, 1, "Catatan", header_format)
    worksheet.write(notes_row + 1, 1, ")* Pemakaian gas CNG mendapatkan fasilitas beban PPN", normal_format)
    worksheet.write(notes_row + 2, 1, ")** Nilai yang harus dibayarkan adalah dasar pengenaan pajak", normal_format)
    
    worksheet.write(notes_row + 4, 1, "Pembayaran harap transfer ke:", header_format)
    worksheet.write(notes_row + 5, 1, "Bank Utama KC Hulu Hilir A/C", normal_format)
    worksheet.write(notes_row + 6, 1, "(IDR) 123 456 7891", normal_format)
    worksheet.write(notes_row + 7, 1, "A/N: PT ENERGI MULTIGUNA", normal_format)
    worksheet.write(notes_row + 8, 1, "", normal_format)
    
    worksheet.write('G1', "", normal_format)

    print(f"Invoice generated")

